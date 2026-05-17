from __future__ import annotations

import csv
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from pirate_monitor.exporters import STATUS_LABELS
from pirate_monitor.storage import Storage


@dataclass(slots=True)
class AnalyticsSummary:
    authors_count: int
    books_count: int
    active_sites_count: int
    findings_count: int
    confirmed_findings_count: int
    full_publication_count: int
    partial_publication_count: int
    removed_count: int
    reposted_count: int
    unknown_count: int
    top_sites: list[dict[str, object]]
    top_authors: list[dict[str, object]]
    top_books: list[dict[str, object]]


class AnalyticsService:
    def __init__(self, storage: Storage | None = None) -> None:
        self.storage = storage or Storage()
        self.full_label = STATUS_LABELS["full"]
        self.partial_label = STATUS_LABELS["partial"]
        self.removed_label = STATUS_LABELS["removed"]
        self.reposted_label = STATUS_LABELS["reposted"]
        self.unknown_label = STATUS_LABELS["unknown"]
        self.status_labels = STATUS_LABELS

    def build_summary(
        self,
        *,
        official_source: str | None = None,
        author_name: str | None = None,
        book_title: str | None = None,
    ) -> AnalyticsSummary:
        rows = self.storage.load_history_rows(
            official_source=official_source,
            author_name=author_name,
            book_title=book_title,
        )
        authors = {
            str(row["source_author"] or "").strip()
            for row in rows
            if row.get("source_author")
        }
        books = {
            (str(row["source_author"] or "").strip(), str(row["source_title"] or "").strip())
            for row in rows
            if row.get("source_author") and row.get("source_title")
        }
        active_sites = {
            str(row["site_name"] or "").strip()
            for row in rows
            if row.get("site_name")
        }
        status_counter = Counter(
            self._normalize_status(row.get("assigned_status")) for row in rows
        )

        per_site: dict[str, list[dict[str, object]]] = defaultdict(list)
        per_author: dict[str, list[dict[str, object]]] = defaultdict(list)
        per_book: dict[tuple[str, str], list[dict[str, object]]] = defaultdict(list)
        confirmed_count = 0
        for row in rows:
            site_name = str(row.get("site_name") or "").strip()
            source_author = str(row.get("source_author") or "").strip()
            source_title = str(row.get("source_title") or "").strip()
            if site_name:
                per_site[site_name].append(row)
            if source_author:
                per_author[source_author].append(row)
            if source_author and source_title:
                per_book[(source_author, source_title)].append(row)
            if bool(row.get("publication_confirmed")):
                confirmed_count += 1

        return AnalyticsSummary(
            authors_count=len(authors),
            books_count=len(books),
            active_sites_count=len(active_sites),
            findings_count=len(rows),
            confirmed_findings_count=confirmed_count,
            full_publication_count=status_counter.get(self.full_label, 0),
            partial_publication_count=status_counter.get(self.partial_label, 0),
            removed_count=status_counter.get(self.removed_label, 0),
            reposted_count=status_counter.get(self.reposted_label, 0),
            unknown_count=status_counter.get(self.unknown_label, 0)
            + status_counter.get("unknown", 0),
            top_sites=self._rank_sites(per_site),
            top_authors=self._rank_authors(per_author),
            top_books=self._rank_books(per_book),
        )

    def export_summary_rows(
        self,
        *,
        official_source: str | None = None,
        author_name: str | None = None,
        book_title: str | None = None,
    ) -> dict[str, list[dict[str, object]]]:
        summary = self.build_summary(
            official_source=official_source,
            author_name=author_name,
            book_title=book_title,
        )
        return {
            "overview": [
                {"metric": "authors_count", "value": summary.authors_count},
                {"metric": "books_count", "value": summary.books_count},
                {"metric": "active_sites_count", "value": summary.active_sites_count},
                {"metric": "findings_count", "value": summary.findings_count},
                {
                    "metric": "confirmed_findings_count",
                    "value": summary.confirmed_findings_count,
                },
                {
                    "metric": "full_publication_count",
                    "value": summary.full_publication_count,
                },
                {
                    "metric": "partial_publication_count",
                    "value": summary.partial_publication_count,
                },
                {"metric": "removed_count", "value": summary.removed_count},
                {"metric": "reposted_count", "value": summary.reposted_count},
                {"metric": "unknown_count", "value": summary.unknown_count},
            ],
            "top_sites": summary.top_sites,
            "top_authors": summary.top_authors,
            "top_books": summary.top_books,
        }

    def export_report(
        self,
        *,
        output_dir: str | Path,
        official_source: str | None = None,
        author_name: str | None = None,
        book_title: str | None = None,
        stamp: str,
    ) -> dict[str, object]:
        tables = self.export_summary_rows(
            official_source=official_source,
            author_name=author_name,
            book_title=book_title,
        )
        target_dir = self._ensure_writable_dir(Path(output_dir))
        slug_parts = [
            (author_name or "").strip(),
            (book_title or "").strip(),
            (official_source or "").strip(),
            stamp,
        ]
        slug = (
            "_".join(part.casefold().replace(" ", "_") for part in slug_parts if part)
            or "analytics"
        )
        paths: dict[str, Path] = {}
        for table_name, rows in tables.items():
            csv_path = self._unique_output_path(target_dir / f"{table_name}_{slug}.csv")
            xlsx_path = self._unique_output_path(target_dir / f"{table_name}_{slug}.xlsx")
            self._export_table_csv(rows, csv_path)
            self._export_table_xlsx(table_name, rows, xlsx_path)
            paths[f"{table_name}_csv"] = csv_path.resolve()
            paths[f"{table_name}_xlsx"] = xlsx_path.resolve()
        return {"tables": tables, **paths}

    def _rank_sites(self, grouped: dict[str, list[dict[str, object]]]) -> list[dict[str, object]]:
        ranked = []
        for site_name, rows in grouped.items():
            statuses = Counter(self._normalize_status(row.get("assigned_status")) for row in rows)
            ranked.append(
                {
                    "site_name": site_name,
                    "findings_count": len(rows),
                    "confirmed_findings_count": sum(
                        1 for row in rows if bool(row.get("publication_confirmed"))
                    ),
                    "full_publication_count": statuses.get(self.full_label, 0),
                    "partial_publication_count": statuses.get(self.partial_label, 0),
                    "removed_count": statuses.get(self.removed_label, 0),
                    "reposted_count": statuses.get(self.reposted_label, 0),
                }
            )
        return sorted(
            ranked,
            key=lambda item: (int(item["findings_count"]), int(item["confirmed_findings_count"])),
            reverse=True,
        )

    def _rank_authors(self, grouped: dict[str, list[dict[str, object]]]) -> list[dict[str, object]]:
        ranked = []
        for author_name, rows in grouped.items():
            books = {
                (
                    str(row.get("source_author") or "").strip(),
                    str(row.get("source_title") or "").strip(),
                )
                for row in rows
            }
            statuses = Counter(self._normalize_status(row.get("assigned_status")) for row in rows)
            ranked.append(
                {
                    "author_name": author_name,
                    "books_count": len(books),
                    "findings_count": len(rows),
                    "active_sites_count": len(
                        {
                            str(row.get("site_name") or "").strip()
                            for row in rows
                            if row.get("site_name")
                        }
                    ),
                    "partial_publication_count": statuses.get(self.partial_label, 0),
                    "full_publication_count": statuses.get(self.full_label, 0),
                    "removed_count": statuses.get(self.removed_label, 0),
                }
            )
        return sorted(
            ranked,
            key=lambda item: (int(item["findings_count"]), int(item["books_count"])),
            reverse=True,
        )

    def _rank_books(
        self, grouped: dict[tuple[str, str], list[dict[str, object]]]
    ) -> list[dict[str, object]]:
        ranked = []
        for (author_name, title), rows in grouped.items():
            statuses = Counter(self._normalize_status(row.get("assigned_status")) for row in rows)
            ranked.append(
                {
                    "author_name": author_name,
                    "title": title,
                    "findings_count": len(rows),
                    "active_sites_count": len(
                        {
                            str(row.get("site_name") or "").strip()
                            for row in rows
                            if row.get("site_name")
                        }
                    ),
                    "top_status": statuses.most_common(1)[0][0]
                    if statuses
                    else self.unknown_label,
                    "confirmed_findings_count": sum(
                        1 for row in rows if bool(row.get("publication_confirmed"))
                    ),
                }
            )
        return sorted(
            ranked,
            key=lambda item: (int(item["findings_count"]), int(item["active_sites_count"])),
            reverse=True,
        )

    def _normalize_status(self, value: object) -> str:
        raw = str(value or "").strip()
        if not raw:
            return self.unknown_label
        return self.status_labels.get(raw, raw)

    def _export_table_csv(self, rows: list[dict[str, object]], output_path: Path) -> None:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        headers = list(rows[0].keys()) if rows else ["metric", "value"]
        with output_path.open("w", encoding="utf-8-sig", newline="") as fh:
            writer = csv.DictWriter(fh, fieldnames=headers)
            writer.writeheader()
            for row in rows:
                writer.writerow({key: row.get(key, "") for key in headers})

    def _export_table_xlsx(
        self, sheet_title: str, rows: list[dict[str, object]], output_path: Path
    ) -> None:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        headers = list(rows[0].keys()) if rows else ["metric", "value"]
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = sheet_title[:31] or "analytics"
        sheet.append(headers)
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        for row in rows:
            sheet.append([row.get(header, "") for header in headers])
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for index, header in enumerate(headers, start=1):
            column = sheet.cell(row=1, column=index).column_letter
            width = max(len(str(header)) + 2, 18)
            for row_index in range(2, sheet.max_row + 1):
                width = min(
                    max(width, len(str(sheet.cell(row=row_index, column=index).value or "")) + 2),
                    40,
                )
                sheet.cell(row=row_index, column=index).alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                )
            sheet.column_dimensions[column].width = width
        workbook.save(output_path)

    def _unique_output_path(self, output_path: Path) -> Path:
        if not output_path.exists():
            return output_path
        stem = output_path.stem
        suffix = output_path.suffix
        parent = output_path.parent
        for index in range(1, 1000):
            candidate = parent / f"{stem}_{index}{suffix}"
            if not candidate.exists():
                return candidate
        raise RuntimeError(f"Не удалось подобрать уникальное имя файла для {output_path}")

    def _ensure_writable_dir(self, output_dir: Path) -> Path:
        try:
            output_dir.mkdir(parents=True, exist_ok=True)
            probe = output_dir / ".write_test"
            probe.write_text("ok", encoding="utf-8")
            probe.unlink(missing_ok=True)
            return output_dir
        except PermissionError:
            fallback = Path("reports/analytics")
            fallback.mkdir(parents=True, exist_ok=True)
            return fallback
