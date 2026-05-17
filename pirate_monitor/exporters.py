from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from pirate_monitor.models import PirateRecord


SHEET_TITLE = "Результаты"
HISTORY_SHEET_TITLE = "История"
EXPORT_HEADERS = [
    "Автор",
    "Произведение",
    "Официальный источник",
    "Ссылка на официальное произведение",
    "Объём текста на официальном сайте",
    "Объём текста на сайте",
    "Подтверждение публикации",
    "Ссылка на подтвержденный текст/файл",
    "Пиратский сайт",
    "Название на сайте",
    "Ссылка на найденную копию",
    "Последнее обновление на сайте",
    "Статус",
]
HISTORY_HEADERS = ["Дата проверки", *EXPORT_HEADERS]
STATUS_LABELS = {
    "full": "опубликован полностью",
    "partial": "опубликован частично",
    "removed": "удален",
    "reposted": "опубликован повторно",
    "unknown": "не определен",
}
OFFICIAL_SOURCE_LABELS = {
    "litnet": "Литнет",
    "author_today": "Автор.Тудей",
    "litres": "ЛитРес",
    "litgorod": "Литгород",
    "litmarket": "Литмаркет",
}


def export_csv(records: list[PirateRecord], output_path: str | Path) -> str:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.writer(fh)
        writer.writerow(EXPORT_HEADERS)
        for record in records:
            writer.writerow(_record_to_row(record))
    return str(output_path.resolve())


def export_xlsx(records: list[PirateRecord], output_path: str | Path) -> str:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = SHEET_TITLE
    sheet.append(EXPORT_HEADERS)

    header_alignment = Alignment(vertical="top", wrap_text=True)
    default_alignment = Alignment(vertical="top")
    wrapped_alignment = Alignment(vertical="top", wrap_text=True)

    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = header_alignment

    for row in map(_record_to_row, records):
        sheet.append(row)

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    widths = {
        "A": 22,
        "B": 34,
        "C": 18,
        "D": 46,
        "E": 22,
        "F": 22,
        "G": 32,
        "H": 44,
        "I": 20,
        "J": 34,
        "K": 52,
        "L": 24,
        "M": 24,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    wrapped_columns = {"B", "D", "E", "F", "G", "H", "J", "K", "L", "M"}
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrapped_alignment if cell.column_letter in wrapped_columns else default_alignment

    workbook.save(output_path)
    return str(output_path.resolve())


def export_history_csv(rows: list[dict[str, object]], output_path: str | Path) -> str:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.writer(fh)
        writer.writerow(HISTORY_HEADERS)
        for row in rows:
            writer.writerow(_history_row_to_values(row))
    return str(output_path.resolve())


def export_history_xlsx(rows: list[dict[str, object]], output_path: str | Path) -> str:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = HISTORY_SHEET_TITLE
    sheet.append(HISTORY_HEADERS)

    header_alignment = Alignment(vertical="top", wrap_text=True)
    default_alignment = Alignment(vertical="top")
    wrapped_alignment = Alignment(vertical="top", wrap_text=True)

    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = header_alignment

    for row in rows:
        sheet.append(_history_row_to_values(row))

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    widths = {
        "A": 22,
        "B": 22,
        "C": 34,
        "D": 18,
        "E": 46,
        "F": 22,
        "G": 22,
        "H": 32,
        "I": 44,
        "J": 20,
        "K": 34,
        "L": 52,
        "M": 24,
        "N": 24,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    wrapped_columns = {"A", "C", "E", "F", "G", "H", "I", "K", "L", "M", "N"}
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrapped_alignment if cell.column_letter in wrapped_columns else default_alignment

    workbook.save(output_path)
    return str(output_path.resolve())


def _record_to_row(record: PirateRecord) -> list[object]:
    return [
        record.source_author or "",
        record.source_title or "",
        OFFICIAL_SOURCE_LABELS.get(record.official_source, record.official_source or ""),
        record.official_url or "",
        _format_official_volume(record),
        _format_pirate_volume(record),
        _format_publication_confirmation(record),
        record.publication_url or "",
        record.site_name or "",
        record.page_title or "",
        record.book_url or "",
        record.last_update_raw or "",
        STATUS_LABELS.get(record.assigned_status, record.assigned_status or "не определен"),
    ]


def _format_official_volume(record: PirateRecord) -> str:
    if record.official_page_count:
        return f"{record.official_page_count} стр."
    if record.official_char_count:
        return f"{record.official_char_count} знаков"
    return ""


def _format_pirate_volume(record: PirateRecord) -> str:
    if record.char_count:
        return f"{record.char_count} знаков"
    return ""


def _format_publication_confirmation(record: PirateRecord) -> str:
    if record.publication_confirmed and record.publication_source == "download":
        file_format = record.publication_format or "файл"
        return f"подтверждено: скачиваемый файл книги ({file_format})"
    if record.publication_confirmed and record.publication_source == "inline":
        return "подтверждено: текст книги на странице сайта"
    if record.publication_source == "card":
        return "не подтверждено: только карточка/аннотация"
    return "не подтверждено"


def _history_row_to_values(row: dict[str, object]) -> list[object]:
    return [
        row.get("checked_at") or "",
        row.get("source_author") or "",
        row.get("source_title") or "",
        OFFICIAL_SOURCE_LABELS.get(str(row.get("official_source") or ""), str(row.get("official_source") or "")),
        row.get("official_url") or "",
        _format_history_official_volume(row),
        _format_history_pirate_volume(row),
        _format_history_publication_confirmation(row),
        row.get("publication_url") or "",
        row.get("site_name") or "",
        row.get("page_title") or "",
        row.get("book_url") or "",
        row.get("last_update_raw") or "",
        STATUS_LABELS.get(str(row.get("assigned_status") or ""), str(row.get("assigned_status") or "не определен")),
    ]


def _format_history_official_volume(row: dict[str, object]) -> str:
    page_count = row.get("official_page_count")
    char_count = row.get("official_char_count")
    if page_count:
        return f"{page_count} стр."
    if char_count:
        return f"{char_count} знаков"
    return ""


def _format_history_pirate_volume(row: dict[str, object]) -> str:
    char_count = row.get("char_count")
    if char_count:
        return f"{char_count} знаков"
    return ""


def _format_history_publication_confirmation(row: dict[str, object]) -> str:
    confirmed = bool(row.get("publication_confirmed"))
    source = str(row.get("publication_source") or "")
    file_format = str(row.get("publication_format") or "")
    if confirmed and source == "download":
        return f"подтверждено: скачиваемый файл книги ({file_format or 'файл'})"
    if confirmed and source == "inline":
        return "подтверждено: текст книги на странице сайта"
    if source == "card":
        return "не подтверждено: только карточка/аннотация"
    return "не подтверждено"
