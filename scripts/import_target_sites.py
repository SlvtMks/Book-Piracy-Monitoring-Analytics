from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook


def main() -> int:
    source_path = Path(r"C:/Users/slvtm/Desktop/Список пиратских сайтов.xlsx")
    output_path = Path("config/target_sites.imported.json")
    workbook = load_workbook(source_path, data_only=True)
    sheet = workbook.active

    rows = []
    for row in sheet.iter_rows(min_row=3, values_only=True):
        name = row[1]
        site = row[2]
        complaint_format = row[3]
        if not name or not site:
            continue
        rows.append(
            {
                "name": str(name).strip(),
                "base_url": str(site).strip(),
                "complaint_format": str(complaint_format or "").strip(),
                "parser": "generic",
                "enabled": "недоступно" not in str(complaint_format or "").casefold()
                and "незачем" not in str(complaint_format or "").casefold(),
            }
        )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")
    print(output_path.resolve())
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
